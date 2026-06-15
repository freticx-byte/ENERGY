import asyncio
import os
import smtplib
from email.message import EmailMessage
from datetime import datetime
from aiogram import Bot, Dispatcher, types
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters import Command
from aiogram.dispatcher.filters.state import State, StatesGroup
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.utils import executor
from openpyxl import Workbook, load_workbook
from apscheduler.schedulers.asyncio import AsyncIOScheduler

TOKEN = "8976307638:AAGZNiGdfhYeYTjWVvWS2g3bAFM5RiwLi1g"

storage = MemoryStorage()
bot = Bot(token=TOKEN)
dp = Dispatcher(bot, storage=storage)

EXCEL_ELE_FILE = "Energy_rep_ELE.xlsx"
EXCEL_RES_FILE = "Energy_rep_RES.xlsx"

EMAIL_TO = "a.misyunas@uvelka.ru"
EMAIL_FROM = "uvenergorusursy@gmail.com"
EMAIL_PASSWORD = "bmdt pzqh qdme wgnc"
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 465

# ============ МЕНЮ ПЛК (2 УРОВНЯ: ГРУППА → СЧЁТЧИК) ============
ELE_MENU = {
    "ЦРП 10кВ": ["ЦРП В1", "ЦРП В2", "ЦРП ТП1, СШ1", "ЦРП ТП1, СШ2",
                 "ЦРП ТП2, СШ1", "ЦРП ТП2, СШ2", "ЦРП ТП3, СШ1", "ЦРП ТП3, СШ2",
                 'ЦРП КСО "Радуга"'],
    "ТП1": ["ТП1 ГРЩ. ППУ В1", "ТП1 ГРЩ. ППУ В2", "ТП1 СГП В1", "ТП1 СГП В2"],
    "ТП2": ["ТП2 ГРЩ В1", "ТП2 ГРЩ В2", "ТП2 Цэх1 В1", "ТП2 Цэх1 В2"],
    "ТП3": [
        "ТП3 Элеватор В1", "ТП3 Элеватор В2",
        "ТП3 Лузговая В1", "ТП3 Лузговая В2",
        "ТП3 ЛОС В1", "ТП3 ЛОС В2",
        "ТП3 КНС В1", "ТП3 КНС В2",
        "ТП3 Насосная В1", "ТП3 Насосная В2",
        "ТП3 Газовая котельная В1", "ТП3 Газовая котельная В2"
    ],
    "ТП4": ["ТП4 Элеватор В1", "ТП4 Элеватор В2", "ТП4 ЛОС В1", "ТП4 ЛОС В2"],
    "ТП5": ["ТП5 ССиТ В1", "ТП5 ССиТ В2"],
    "ТП6": ["ТП6 ГЦ В1", "ТП6 ГЦ В2"],
    "ТП7": ["ТП7 СТЗ"],
    "ГПУ КТПН": [
        "КТПН ГПУ →", "КТПН ГПУ ←",
        "КТПН ГПУ1 →", "КТПН ГПУ1 ←",
        "КТПН ГПУ2 →", "КТПН ГПУ2 ←",
        "КТПН ГПУ3 →", "КТПН ГПУ3 ←",
        "КТПН ГПУ4 →", "КТПН ГПУ4 ←"
    ],
    "Радуга Склад": ["ШВ-6", "ШС-6", "ЩС Зар В1", "ЩС Зар В2"],
    "Теплосети": ["Теплосети ИТП"],
    "Прочее": ["ТП Луговская", "Склад газации", "Временно ТП-3"]
}

# ============ МЕНЮ РЕСУРС (2 УРОВНЯ: ГРУППА → СЧЁТЧИК) ============
RES_MENU = {
    "Цех1": ["РП3"],
    "Цех3": ["ВРУ-1 (ТП-304)", "ВРУ-1 (ТП-324)", "ВРУ-2 (ТП-304)", "ВРУ-2 (ТП-324)"],
    "Цех4": ["ТЛ1 ШУ-3-1 ШУ6", "ТЛ2 ШУ-3-2 ШУ6", "ТЛ3 ШУ-3-1 ШУ5", "Нории14-23 ШУ5"],
    "РМЦ": ["РМЦ", "Карный участок"],
    "Компрессорная": [
        "ПК3 ТП304 В1", "ПК3 ТП324 В2",
        "ПК2 Компр.4", "ПК2 Компр.5", "ПК2 Компр.6",
        "ПК1 Компр.7", "ПК1 Компр.8"
    ],
    "РХУ": ["Скл. бестарного хранения", "Хлопья Шулле"],
    "Фасовка": ["Фасовка", "ЩВ фасовка"],
    "ССиТ": ["ВРУ4 ТП304 В1", "ВРУ4 ТП324 В2"],
    "Офис": ["Офис (Админ)"],
    "Столовая": ["ШР столовая"],
    "КГУ": ["Выработка КГУ"]
}

# Плоские списки счётчиков
ALL_ELE_COUNTERS = []
for counters in ELE_MENU.values():
    ALL_ELE_COUNTERS.extend(counters)

ALL_RES_COUNTERS = []
for counters in RES_MENU.values():
    ALL_RES_COUNTERS.extend(counters)


class EnergyForm(StatesGroup):
    choosing_object = State()
    choosing_group = State()
    entering_value = State()


# ============ ФУНКЦИИ EXCEL ============
def init_excel(file_path, counters):
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.append(["Дата", "Время записи"] + counters)
        wb.save(file_path)
        print(f"✅ Создан файл {file_path}")


def get_today_str():
    return datetime.now().strftime("%Y-%m-%d")


def ensure_today_exists(file_path, counters):
    today = get_today_str()
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        date_exists = False
        for row in range(2, ws.max_row + 1):
            if ws.cell(row, 1).value == today:
                date_exists = True
                break
        if not date_exists:
            ws.append([today, ""] + [0] * len(counters))
            wb.save(file_path)
        wb.close()
    except:
        pass


def update_reading(file_path, counter_name, value, record_time, counters):
    try:
        today = get_today_str()
        ensure_today_exists(file_path, counters)
        
        wb = load_workbook(file_path)
        ws = wb.active
        headers = [str(cell.value) if cell.value else "" for cell in ws[1]]
        
        if counter_name not in headers:
            wb.close()
            return False
        
        col_idx = headers.index(counter_name) + 1
        
        for row in range(2, ws.max_row + 1):
            if ws.cell(row, 1).value == today:
                ws.cell(row, col_idx).value = float(value)
                ws.cell(row, 2).value = record_time
                wb.save(file_path)
                wb.close()
                return True
        wb.close()
        return False
    except:
        return False


def get_all_data(file_path):
    try:
        if not os.path.exists(file_path):
            return []
        wb = load_workbook(file_path)
        ws = wb.active
        result = []
        for row in range(2, ws.max_row + 1):
            date = ws.cell(row, 1).value
            if not date:
                continue
            total = 0
            for col in range(3, ws.max_column + 1):
                val = ws.cell(row, col).value
                if val and isinstance(val, (int, float)) and val != 0:
                    total += val
            if total > 0:
                result.append({"date": date, "total": total})
        wb.close()
        return result
    except:
        return []


def get_counters_with_data(file_path):
    try:
        if not os.path.exists(file_path):
            return []
        wb = load_workbook(file_path)
        ws = wb.active
        headers = [str(cell.value) if cell.value else "" for cell in ws[1]][2:]
        counters_with_data = set()
        for row in range(2, ws.max_row + 1):
            for idx, counter in enumerate(headers, start=3):
                val = ws.cell(row, idx).value
                if val and isinstance(val, (int, float)) and val != 0:
                    counters_with_data.add(counter)
        wb.close()
        return list(counters_with_data)
    except:
        return []


# ============ EMAIL ============
async def send_email_report(file_path, name):
    if not os.path.exists(file_path):
        return False
    try:
        msg = EmailMessage()
        msg['Subject'] = f"Отчёт по {name} за {datetime.now().strftime('%d.%m.%Y')}"
        msg['From'] = EMAIL_FROM
        msg['To'] = EMAIL_TO
        msg.set_content(f"Отчёт по {name}\nДата: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
        with open(file_path, 'rb') as f:
            msg.add_attachment(f.read(), maintype='application',
                               subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                               filename=f"{name}_{datetime.now().strftime('%Y%m%d')}.xlsx")
        with smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT) as server:
            server.login(EMAIL_FROM, EMAIL_PASSWORD)
            server.send_message(msg)
        return True
    except:
        return False


# ============ КЛАВИАТУРЫ ============
def get_main_menu():
    keyboard = ReplyKeyboardMarkup(resize_keyboard=True)
    keyboard.add(KeyboardButton("📊 ЭЭ ПЛК"))
    keyboard.add(KeyboardButton("📊 ЭЭ Ресурс"))
    keyboard.add(KeyboardButton("📁 Скачать Excel"))
    keyboard.add(KeyboardButton("❓ Помощь"))
    return keyboard


def get_object_keyboard():
    keyboard = InlineKeyboardMarkup(row_width=2)
    keyboard.add(InlineKeyboardButton("📊 ЭЭ ПЛК", callback_data="obj_ELE"))
    keyboard.add(InlineKeyboardButton("📊 ЭЭ Ресурс", callback_data="obj_RES"))
    keyboard.add(InlineKeyboardButton("❌ Отмена", callback_data="cancel"))
    return keyboard


def get_group_keyboard(menu):
    keyboard = InlineKeyboardMarkup(row_width=1)
    for group_name in menu.keys():
        keyboard.add(InlineKeyboardButton(group_name, callback_data=f"group_{group_name}"))
    keyboard.add(InlineKeyboardButton("❌ Отмена", callback_data="cancel"))
    return keyboard


def get_counters_keyboard(group_name, menu):
    keyboard = InlineKeyboardMarkup(row_width=1)
    for counter in menu[group_name]:
        text = counter[:40] if len(counter) > 40 else counter
        keyboard.add(InlineKeyboardButton(text, callback_data=f"cnt_{counter}"))
    keyboard.add(InlineKeyboardButton("🔙 Назад к группам", callback_data="back_to_groups"))
    keyboard.add(InlineKeyboardButton("❌ Отмена", callback_data="cancel"))
    return keyboard


# ============ ОБРАБОТЧИКИ ============
@dp.message_handler(commands=["start"])
async def start(message: types.Message):
    init_excel(EXCEL_ELE_FILE, ALL_ELE_COUNTERS)
    init_excel(EXCEL_RES_FILE, ALL_RES_COUNTERS)
    
    data_ele = get_all_data(EXCEL_ELE_FILE)
    data_res = get_all_data(EXCEL_RES_FILE)
    
    text = (
        f"🏭 ЭНЕРГОУЧЁТ\n\n"
        f"📅 Сегодня: {datetime.now().strftime('%d.%m.%Y')}\n"
        f"📊 ЭЭ ПЛК: {len(data_ele)} записей, {len(ALL_ELE_COUNTERS)} счётчиков\n"
        f"📊 ЭЭ Ресурс: {len(data_res)} записей, {len(ALL_RES_COUNTERS)} счётчиков\n\n"
        f"💡 Нажмите '📊 ЭЭ ПЛК' или '📊 ЭЭ Ресурс'"
    )
    await message.answer(text, reply_markup=get_main_menu())


@dp.message_handler(lambda message: message.text == "📊 ЭЭ ПЛК")
async def add_ele(message: types.Message, state: FSMContext):
    await state.update_data(object_type="ELE", file_path=EXCEL_ELE_FILE, counters=ALL_ELE_COUNTERS, menu=ELE_MENU)
    await message.answer("📁 Выберите группу ПЛК:", reply_markup=get_group_keyboard(ELE_MENU))
    await state.set_state(EnergyForm.choosing_group)


@dp.message_handler(lambda message: message.text == "📊 ЭЭ Ресурс")
async def add_res(message: types.Message, state: FSMContext):
    await state.update_data(object_type="RES", file_path=EXCEL_RES_FILE, counters=ALL_RES_COUNTERS, menu=RES_MENU)
    await message.answer("📁 Выберите группу Ресурс:", reply_markup=get_group_keyboard(RES_MENU))
    await state.set_state(EnergyForm.choosing_group)


@dp.callback_query_handler(lambda c: c.data.startswith("group_"), state=EnergyForm.choosing_group)
async def group_selected(callback_query: types.CallbackQuery, state: FSMContext):
    group_name = callback_query.data.replace("group_", "")
    data = await state.get_data()
    menu = data.get('menu')
    await state.update_data(selected_group=group_name)
    await callback_query.message.edit_text(
        f"📁 {group_name}\n\n👇 Выберите счётчик:",
        reply_markup=get_counters_keyboard(group_name, menu)
    )
    await callback_query.answer()


@dp.callback_query_handler(lambda c: c.data == "back_to_groups", state=EnergyForm.choosing_group)
async def back_to_groups(callback_query: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    menu = data.get('menu')
    obj = data.get('object_type', 'ELE')
    title = "ПЛК" if obj == "ELE" else "Ресурс"
    await callback_query.message.edit_text(
        f"📁 Выберите группу {title}:",
        reply_markup=get_group_keyboard(menu)
    )
    await callback_query.answer()


@dp.callback_query_handler(lambda c: c.data.startswith("cnt_"), state=EnergyForm.choosing_group)
async def counter_selected(callback_query: types.CallbackQuery, state: FSMContext):
    counter_name = callback_query.data.replace("cnt_", "")
    await state.update_data(selected_counter=counter_name)
    today = datetime.now().strftime("%d.%m.%Y")
    await callback_query.message.edit_text(
        f"✅ Выбран счётчик: {counter_name}\n\n"
        f"📅 Дата: {today}\n"
        f"✏️ Введите показание (кВт·ч):"
    )
    await state.set_state(EnergyForm.entering_value)
    await callback_query.answer()


@dp.message_handler(state=EnergyForm.entering_value)
async def value_entered(message: types.Message, state: FSMContext):
    try:
        value = float(message.text.replace(",", "."))
        if value < 0:
            await message.answer("❌ Введите положительное число")
            return
        
        data = await state.get_data()
        counter = data['selected_counter']
        group = data['selected_group']
        file_path = data['file_path']
        counters = data['counters']
        menu = data.get('menu')
        record_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        if update_reading(file_path, counter, value, record_time, counters):
            await message.answer(f"✅ Сохранено!\n\n🏭 {counter}\n⚡ {value:,.2f} кВт·ч")
            await message.answer(
                f"📁 {group}\n\n👇 Выберите следующий счётчик:",
                reply_markup=get_counters_keyboard(group, menu)
            )
            await state.set_state(EnergyForm.choosing_group)
        else:
            await message.answer("❌ Ошибка при сохранении!", reply_markup=get_main_menu())
            await state.finish()
        
    except ValueError:
        await message.answer("❌ Введите число (например: 125.5)")


@dp.message_handler(commands=["stats"])
async def show_stats(message: types.Message):
    data_ele = get_all_data(EXCEL_ELE_FILE)
    data_res = get_all_data(EXCEL_RES_FILE)
    
    total_ele = sum(d['total'] for d in data_ele)
    total_res = sum(d['total'] for d in data_res)
    
    text = (
        f"📊 СТАТИСТИКА\n\n"
        f"📊 ЭЭ ПЛК:\n"
        f"   Записей: {len(data_ele)}\n"
        f"   Сумма: {total_ele:,.2f} кВт·ч\n\n"
        f"📊 ЭЭ Ресурс:\n"
        f"   Записей: {len(data_res)}\n"
        f"   Сумма: {total_res:,.2f} кВт·ч"
    )
    await message.answer(text, reply_markup=get_main_menu())


@dp.message_handler(lambda message: message.text == "📁 Скачать Excel")
@dp.message_handler(commands=["file"])
async def send_excel_file(message: types.Message):
    keyboard = InlineKeyboardMarkup(row_width=2)
    keyboard.add(InlineKeyboardButton("📊 ЭЭ ПЛК", callback_data="download_ELE"))
    keyboard.add(InlineKeyboardButton("📊 ЭЭ Ресурс", callback_data="download_RES"))
    await message.answer("Выберите файл:", reply_markup=keyboard)


@dp.callback_query_handler(lambda c: c.data.startswith("download_"))
async def download_file(callback_query: types.CallbackQuery):
    obj = callback_query.data.replace("download_", "")
    file_path = EXCEL_ELE_FILE if obj == "ELE" else EXCEL_RES_FILE
    name = "ЭЭ ПЛК" if obj == "ELE" else "ЭЭ Ресурс"
    
    if os.path.exists(file_path):
        doc = types.InputFile(file_path)
        await bot.send_document(callback_query.from_user.id, doc, caption=f"📊 {name}\n{datetime.now().strftime('%d.%m.%Y %H:%M')}")
    else:
        await bot.send_message(callback_query.from_user.id, "❌ Файл не создан")
    await callback_query.answer()


@dp.message_handler(commands=["test_email"])
async def test_email(message: types.Message):
    await message.answer("📧 Отправляю...")
    r1 = await send_email_report(EXCEL_ELE_FILE, "ЭЭ_ПЛК")
    r2 = await send_email_report(EXCEL_RES_FILE, "ЭЭ_Ресурс")
    await message.answer(f"✅ ПЛК: {r1}, Ресурс: {r2}")


@dp.message_handler(commands=["help"])
@dp.message_handler(lambda message: message.text == "❓ Помощь")
async def help_command(message: types.Message):
    text = (
        "📘 ПОМОЩЬ\n\n"
        "📊 ЭЭ ПЛК - Ввод показаний ПЛК\n"
        "📊 ЭЭ Ресурс - Ввод показаний Ресурс\n"
        "📁 Скачать Excel - Скачать файл\n"
        "/stats - Общая статистика\n"
        "/test_email - Проверить почту\n\n"
        f"📊 ПЛК: {len(ALL_ELE_COUNTERS)} счётчиков\n"
        f"📊 Ресурс: {len(ALL_RES_COUNTERS)} счётчиков"
    )
    await message.answer(text, reply_markup=get_main_menu())


@dp.callback_query_handler(lambda c: c.data == "cancel", state="*")
async def cancel_action(callback_query: types.CallbackQuery, state: FSMContext):
    await state.finish()
    await callback_query.message.delete()
    await callback_query.message.answer("❌ Отменено", reply_markup=get_main_menu())
    await callback_query.answer()


if __name__ == "__main__":
    print("🚀 Бот Энергоучёт запущен!")
    print(f"📊 ПЛК: {len(ALL_ELE_COUNTERS)} счётчиков")
    print(f"📊 Ресурс: {len(ALL_RES_COUNTERS)} счётчиков")
    
    init_excel(EXCEL_ELE_FILE, ALL_ELE_COUNTERS)
    init_excel(EXCEL_RES_FILE, ALL_RES_COUNTERS)
    
    scheduler = AsyncIOScheduler(timezone="Europe/Moscow")
    scheduler.add_job(lambda: send_email_report(EXCEL_ELE_FILE, "ЭЭ_ПЛК"), 'cron', day_of_week='mon', hour=12, minute=0)
    scheduler.add_job(lambda: send_email_report(EXCEL_RES_FILE, "ЭЭ_Ресурс"), 'cron', day_of_week='mon', hour=12, minute=0)
    scheduler.start()
    print("⏰ Отчёты: каждый понедельник в 12:00 МСК")
    
    executor.start_polling(dp, skip_updates=True)
